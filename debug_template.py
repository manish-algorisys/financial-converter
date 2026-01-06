"""
Quick test to debug column mapping
"""
from pathlib import Path
from openpyxl import load_workbook

template_path = Path("templates/financial_summary_template_column_mapping.xlsx")
wb = load_workbook(template_path)
ws = wb.active

print("Template contents:")
print("=" * 60)

# Check first 5 rows
for row_num in range(1, 6):
    print(f"\nRow {row_num}:")
    for col_idx in range(1, 7):  # Columns A-F
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        cell = ws[f'{col_letter}{row_num}']
        
        # Handle merged cells
        value = cell.value
        if hasattr(cell, 'coordinate'):
            print(f"  {col_letter}{row_num}: {value}")

print("\n" + "=" * 60)
print("Checking period parsing:")

test_headers = [
    "30.06.2025 Q",
    "31.03.2025 Q",
    "30.06.2024 Q",
    "31.03.2025 Y"
]

# Test the period parsing logic
import re

def parse_period_from_header(header_text: str) -> str:
    """Test period parsing"""
    header_text = str(header_text).strip()
    
    # Pattern 1: Direct date format with Q or Y suffix
    match = re.search(r'(\d{2}\.\d{2}\.\d{4})\s*([QY])', header_text, re.IGNORECASE)
    if match:
        date_part = match.group(1)
        suffix = match.group(2).upper()
        if suffix == 'Y':
            return f"{date_part}_Y"
        return date_part
    
    # Pattern 2: Just the date without suffix
    match = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', header_text)
    if match:
        return match.group(1)
    
    return None

for header in test_headers:
    result = parse_period_from_header(header)
    print(f"  '{header}' -> '{result}'")
