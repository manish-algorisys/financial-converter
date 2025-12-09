"""
Excel/CSV Generator for Financial Data
Converts JSON financial data to formatted Excel and CSV files
"""
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv
import json
import uuid

_log = logging.getLogger(__name__)


class FinancialExcelGenerator:
    """Generate Excel and CSV files from financial JSON data."""
    
    # Period mapping: JSON keys to column positions (B=1, C=2, etc.)
    PERIOD_MAPPING = {
        '30.06.2025': ('B', 1, 'Unaudited Q1', '3M-30th Jun 2025'),
        '31.03.2025_Y': ('C', 2, 'FY 2025', '12M'),
        '31.03.2025': ('D', 3, 'Q4', '3M-31st Mar 2025'),
        '31.12.2024': ('E', 4, 'Q3', '3M-31st Dec 2024'),
        '30.09.2024': ('F', 5, 'Q2', '3M-30th Sept 2024'),
        '30.06.2024': ('G', 6, 'Unaudited Q1 FY 2024', '3M-30th Jun 2024'),
        '31.03.2024_Y': ('H', 7, 'FY 2024', '12M'),
        '31.03.2024': ('I', 8, 'Q4 FY 2024', '3M-31st Mar 2024'),
        '31.12.2023': ('J', 9, 'Q3 FY 2024', '3M-31st Dec 2023'),
        '30.09.2023': ('K', 10, 'Q2 FY 2024', '3M-30th Sept 2023'),
        '30.06.2023': ('L', 11, 'Q1 FY 2024', '3M-30th Jun 2023'),
    }
    
    # Key aliases - support both naming conventions
    KEY_ALIASES = {
        'sale_of_products': 'sale_of_goods',
        'sale_of_goods': 'sale_of_goods',
    }
    
    def __init__(self):
        """Initialize the generator."""
        self.data_map = {}
        self.company_name = ""
    
    def _normalize_key(self, key: str) -> str:
        """Normalize key using aliases."""
        return self.KEY_ALIASES.get(key, key)
    
    def _parse_number(self, value: str) -> float:
        """Parse number from string, handling brackets for negatives."""
        if not value or value.strip() == '':
            return 0.0
        
        value = str(value).strip()
        
        # Handle brackets for negatives: (123) = -123
        if value.startswith('(') and value.endswith(')'):
            value = '-' + value[1:-1]
        
        # Remove commas
        value = value.replace(',', '')
        
        try:
            return float(value)
        except (ValueError, TypeError):
            return 0.0
    
    def _format_number(self, value: float, decimal_places: int = 2) -> str:
        """Format number with Indian comma style and brackets for negatives."""
        if value == 0:
            return '-'
        
        # Handle negative numbers
        is_negative = value < 0
        abs_value = abs(value)
        
        # Format with decimal places
        formatted = f"{abs_value:,.{decimal_places}f}"
        
        # Convert to Indian numbering system (lakhs/crores)
        # For simplicity, keeping comma format but can be enhanced
        
        if is_negative:
            return f"({formatted})"
        return formatted
    
    def _build_data_map(self, financial_data: List[Dict]) -> None:
        """Build data map from financial_data array for easy lookup."""
        self.data_map = {}
        
        for item in financial_data:
            key = self._normalize_key(item.get('key', ''))
            values = item.get('values', {})
            
            if key:
                self.data_map[key] = values
    
    def _get_value(self, key: str, period: str) -> float:
        """Get numeric value for a key and period."""
        normalized_key = self._normalize_key(key)
        
        if normalized_key in self.data_map:
            values = self.data_map[normalized_key]
            value_str = values.get(period, '')
            return self._parse_number(value_str)
        
        return 0.0
    
    def _calculate_total(self, keys: List[str], period: str) -> float:
        """Calculate total from multiple keys."""
        total = 0.0
        for key in keys:
            total += self._get_value(key, period)
        return total
    
    def _create_excel_headers(self, ws) -> None:
        """Create Excel headers (rows 1-3)."""
        # Row 1: Company Name
        ws.merge_cells('B1:L1')
        cell = ws['B1']
        cell.value = self.company_name
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Row 2: Period Headers
        ws['A2'] = 'INR Crs'
        for period_key, (col_letter, _, header, _) in self.PERIOD_MAPPING.items():
            cell = ws[f'{col_letter}2']
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
        
        # Row 3: Period Descriptions
        ws['A3'] = 'I. Revenue from operations'
        ws['A3'].font = Font(bold=True)
        ws['A3'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        
        for period_key, (col_letter, _, _, description) in self.PERIOD_MAPPING.items():
            cell = ws[f'{col_letter}3']
            cell.value = description
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    
    def _apply_cell_borders(self, ws, row: int, start_col: str = 'A', end_col: str = 'L') -> None:
        """Apply thin borders to cells in a row."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(ord(start_col), ord(end_col) + 1):
            cell = ws[f'{chr(col)}{row}']
            cell.border = thin_border
    
    def _apply_total_style(self, ws, row: int) -> None:
        """Apply total line styling (bold, double bottom border)."""
        for col in 'ABCDEFGHIJKL':
            cell = ws[f'{col}{row}']
            cell.font = Font(bold=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='double')
            )
    
    def _set_row_data(self, ws, row: int, label: str, key: str, 
                      is_total: bool = False, is_section_header: bool = False) -> None:
        """Set data for a row across all periods."""
        # Set label
        ws[f'A{row}'] = label
        
        if is_section_header:
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
        elif is_total:
            ws[f'A{row}'].font = Font(bold=True)
        
        # Set data for each period
        for period_key, (col_letter, _, _, _) in self.PERIOD_MAPPING.items():
            value = self._get_value(key, period_key)
            cell = ws[f'{col_letter}{row}']
            
            if value != 0:
                cell.value = self._format_number(value)
            else:
                cell.value = '-'
            
            cell.alignment = Alignment(horizontal='right')
            
            if is_total:
                cell.font = Font(bold=True)
        
        # Apply borders
        self._apply_cell_borders(ws, row)
        
        if is_total:
            self._apply_total_style(ws, row)
    
    def generate_excel(self, json_data: Dict, output_path: Path) -> bool:
        """
        Generate Excel file from JSON financial data.
        
        Args:
            json_data: Financial data in JSON format
            output_path: Path to save Excel file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Extract data
            self.company_name = json_data.get('company_name', 'Financial Statement')
            financial_data = json_data.get('financial_data', [])
            
            # Build data map
            self._build_data_map(financial_data)
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Statement"
            
            # Set column widths
            ws.column_dimensions['A'].width = 60
            for col in 'BCDEFGHIJKL':
                ws.column_dimensions[col].width = 15
            
            # Create headers (rows 1-3)
            self._create_excel_headers(ws)
            
            # Row 4-8: Revenue Section
            self._set_row_data(ws, 4, 'Sale of goods / Income from operations Domestic', 'sale_of_goods')
            self._set_row_data(ws, 5, 'Sale Exports', 'export_sales')
            self._set_row_data(ws, 6, 'Revenue from Services', 'service_revenue')
            self._set_row_data(ws, 7, 'Other operating revenues', 'other_operating_revenues')
            self._set_row_data(ws, 8, 'Total Revenue', 'revenue_from_operations', is_total=True)
            
            # Row 9: Other Income
            self._set_row_data(ws, 9, 'II. Other income', 'other_income', is_section_header=True)
            
            # Row 10-13: Total Income Section
            self._set_row_data(ws, 10, 'III. Total Income (I+II)', 'total_income', is_section_header=True)
            ws['A11'] = 'Sale of Goods Growth YOY'
            ws['A12'] = 'Total Revenue Growth YOY'
            ws['A13'] = 'Total Income Growth YOY'
            
            # Row 14: Empty
            
            # Row 15-32: Expenses Section
            ws['A15'] = 'IV. Expenses:'
            ws['A15'].font = Font(bold=True)
            ws['A15'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            self._set_row_data(ws, 16, 'Cost of materials consumed', 'cost_of_materials_consumed')
            self._set_row_data(ws, 17, 'Excise duty', 'excise_duty')
            self._set_row_data(ws, 18, 'Purchases of stock-in-trade', 'purchases_stock_in_trade')
            self._set_row_data(ws, 19, 'Changes in inventories of finished goods, work-in-progress and stock-in-trade', 
                              'changes_in_inventories')
            self._set_row_data(ws, 20, 'Employee benefits expense', 'employee_benefits_expense')
            self._set_row_data(ws, 21, 'Finance costs', 'finance_costs')
            self._set_row_data(ws, 22, 'Depreciation and amortisation expense', 'depreciation_amortisation_expense')
            self._set_row_data(ws, 23, 'Other expenses', 'other_expense')
            self._set_row_data(ws, 24, 'Advertising and promotion', 'advertising_expense')
            ws['A25'] = 'Others'
            self._set_row_data(ws, 26, 'Impairment', 'impairment_losses')
            ws['A27'] = 'Provision for contingencies'
            ws['A28'] = 'Corporate responsibilities'
            self._set_row_data(ws, 29, 'Total expenses', 'total_expenses', is_total=True)
            self._set_row_data(ws, 30, 'PBT before exp items', 'profit_before_exceptional_and_tax')
            # Row 31: Empty
            self._set_row_data(ws, 32, 'Exceptional items Gain/(Loss)', 'exceptional_item_expense')
            
            # Row 33: Empty
            
            # Row 34-35: Profit Before Tax
            self._set_row_data(ws, 34, 'V. Profit before tax (III-IV)', 'profit_before_tax', is_section_header=True)
            ws['A35'] = '%'
            
            # Row 36: Empty
            
            # Row 37-40: Tax Section
            ws['A37'] = 'VI. Tax expense:'
            ws['A37'].font = Font(bold=True)
            ws['A37'].fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
            
            self._set_row_data(ws, 38, '(i) Current tax', 'current_tax')
            self._set_row_data(ws, 39, '(ii) Deferred tax/Income Tax of Prior years', 'deferred_tax')
            self._set_row_data(ws, 40, 'Total Tax', 'total_tax_expense', is_total=True)
            
            # Row 41: Net Profit
            self._set_row_data(ws, 41, 'VII. Profit for the year (V-VI)', 'net_profit', is_section_header=True)
            
            # Row 42: Empty
            
            # Row 43-44: EBITDA
            self._set_row_data(ws, 43, 'EBITDA', 'ebitda')
            ws['A44'] = 'EBITDA Margin'
            
            # Row 45: Empty
            
            # Row 46-47: Growth Metrics
            ws['A46'] = 'Volume Gr%'
            ws['A47'] = 'Price Gr%'
            
            # Apply borders to all used cells
            for row in range(1, 48):
                self._apply_cell_borders(ws, row)
            
            # Save workbook
            wb.save(output_path)
            _log.info(f"Excel file generated: {output_path}")
            return True
            
        except Exception as e:
            _log.error(f"Error generating Excel: {e}", exc_info=True)
            return False
    
    def generate_csv(self, json_data: Dict, output_path: Path) -> bool:
        """
        Generate CSV file from JSON financial data.
        
        Args:
            json_data: Financial data in JSON format
            output_path: Path to save CSV file
        
        Returns:
            True if successful, False otherwise
        """
        try:
            # Extract data
            self.company_name = json_data.get('company_name', 'Financial Statement')
            financial_data = json_data.get('financial_data', [])
            
            # Build data map
            self._build_data_map(financial_data)
            
            # Prepare rows
            rows = []
            
            # Row 1: Company Name
            row1 = [''] + [self.company_name] + [''] * 10
            rows.append(row1)
            
            # Row 2: Period Headers
            row2 = ['INR Crs']
            for period_key, (_, _, header, _) in sorted(self.PERIOD_MAPPING.items(), 
                                                        key=lambda x: x[1][1]):
                row2.append(header)
            rows.append(row2)
            
            # Row 3: Period Descriptions
            row3 = ['I. Revenue from operations']
            for period_key, (_, _, _, description) in sorted(self.PERIOD_MAPPING.items(), 
                                                             key=lambda x: x[1][1]):
                row3.append(description)
            rows.append(row3)
            
            # Helper function to create data row
            def create_row(label: str, key: str) -> List[str]:
                row = [label]
                for period_key, (_, _, _, _) in sorted(self.PERIOD_MAPPING.items(), 
                                                       key=lambda x: x[1][1]):
                    value = self._get_value(key, period_key)
                    if value != 0:
                        row.append(self._format_number(value))
                    else:
                        row.append('-')
                return row
            
            # Add all rows according to structure
            rows.append(create_row('Sale of goods / Income from operations Domestic', 'sale_of_goods'))
            rows.append(create_row('Sale Exports', 'export_sales'))
            rows.append(create_row('Revenue from Services', 'service_revenue'))
            rows.append(create_row('Other operating revenues', 'other_operating_revenues'))
            rows.append(create_row('Total Revenue', 'revenue_from_operations'))
            
            rows.append(create_row('II. Other income', 'other_income'))
            
            rows.append(create_row('III. Total Income (I+II)', 'total_income'))
            rows.append(['Sale of Goods Growth YOY'] + ['-'] * 11)
            rows.append(['Total Revenue Growth YOY'] + ['-'] * 11)
            rows.append(['Total Income Growth YOY'] + ['-'] * 11)
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(['IV. Expenses:'] + ['-'] * 11)
            rows.append(create_row('Cost of materials consumed', 'cost_of_materials_consumed'))
            rows.append(create_row('Excise duty', 'excise_duty'))
            rows.append(create_row('Purchases of stock-in-trade', 'purchases_stock_in_trade'))
            rows.append(create_row('Changes in inventories of finished goods, work-in-progress and stock-in-trade', 
                                  'changes_in_inventories'))
            rows.append(create_row('Employee benefits expense', 'employee_benefits_expense'))
            rows.append(create_row('Finance costs', 'finance_costs'))
            rows.append(create_row('Depreciation and amortisation expense', 'depreciation_amortisation_expense'))
            rows.append(create_row('Other expenses', 'other_expense'))
            rows.append(create_row('Advertising and promotion', 'advertising_expense'))
            rows.append(['Others'] + ['-'] * 11)
            rows.append(create_row('Impairment', 'impairment_losses'))
            rows.append(['Provision for contingencies'] + ['-'] * 11)
            rows.append(['Corporate responsibilities'] + ['-'] * 11)
            rows.append(create_row('Total expenses', 'total_expenses'))
            rows.append(create_row('PBT before exp items', 'profit_before_exceptional_and_tax'))
            rows.append([''] * 12)  # Empty row
            rows.append(create_row('Exceptional items Gain/(Loss)', 'exceptional_item_expense'))
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(create_row('V. Profit before tax (III-IV)', 'profit_before_tax'))
            rows.append(['%'] + ['-'] * 11)
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(['VI. Tax expense:'] + ['-'] * 11)
            rows.append(create_row('(i) Current tax', 'current_tax'))
            rows.append(create_row('(ii) Deferred tax/Income Tax of Prior years', 'deferred_tax'))
            rows.append(create_row('Total Tax', 'total_tax_expense'))
            
            rows.append(create_row('VII. Profit for the year (V-VI)', 'net_profit'))
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(create_row('EBITDA', 'ebitda'))
            rows.append(['EBITDA Margin'] + ['-'] * 11)
            
            rows.append([''] * 12)  # Empty row
            
            rows.append(['Volume Gr%'] + ['-'] * 11)
            rows.append(['Price Gr%'] + ['-'] * 11)
            
            # Write CSV
            with open(output_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(rows)
            
            _log.info(f"CSV file generated: {output_path}")
            return True
            
        except Exception as e:
            _log.error(f"Error generating CSV: {e}", exc_info=True)
            return False


class FileManager:
    """Manage generated Excel/CSV files with metadata."""
    
    def __init__(self, storage_dir: Path):
        """
        Initialize file manager.
        
        Args:
            storage_dir: Directory to store files and metadata
        """
        self.storage_dir = storage_dir
        self.storage_dir.mkdir(parents=True, exist_ok=True)
        self.metadata_file = storage_dir / 'metadata.json'
        self.metadata = self._load_metadata()
    
    def _load_metadata(self) -> Dict:
        """Load metadata from disk."""
        if self.metadata_file.exists():
            try:
                with open(self.metadata_file, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                _log.error(f"Error loading metadata: {e}")
                return {}
        return {}
    
    def _save_metadata(self) -> None:
        """Save metadata to disk."""
        print(f"Saving metadata: {self.metadata}")
        try:
            with open(self.metadata_file, 'w', encoding='utf-8') as f:
                json.dump(self.metadata, f, indent=2)
        except Exception as e:
            _log.error(f"Error saving metadata: {e}")
    
    def save_file(self, file_path: Path, company_name: str, file_type: str) -> str:
        """
        Save file and create metadata entry.
        
        Args:
            file_path: Path to the generated file
            company_name: Company name
            file_type: 'excel' or 'csv'
        
        Returns:
            Unique file ID
        """
        file_id = str(uuid.uuid4())
        timestamp = datetime.now().isoformat()
        file_size = file_path.stat().st_size
        
        # Copy file to storage with unique name
        extension = '.xlsx' if file_type == 'excel' else '.csv'
        stored_path = self.storage_dir / f"{file_id}{extension}"
        print(f"stored_path: {stored_path}")
        
        import shutil
        shutil.copy2(file_path, stored_path)
        
        # Create metadata
        self.metadata[file_id] = {
            'file_id': file_id,
            'company_name': company_name,
            'file_type': file_type,
            'original_name': file_path.name,
            'stored_path': str(stored_path),
            'created_at': timestamp,
            'file_size': file_size,
            'download_count': 0,
            'stored_path': str(stored_path)
        }
        
        self._save_metadata()
        _log.info(f"File saved with ID: {file_id}")
        
        return file_id
    
    def get_file(self, file_id: str) -> Optional[Dict]:
        """
        Get file metadata and increment download count.
        
        Args:
            file_id: File ID
        
        Returns:
            File metadata or None if not found
        """
        if file_id in self.metadata:
            self.metadata[file_id]['download_count'] += 1
            self._save_metadata()
            return self.metadata[file_id]
        return None
    
    def list_files(self, company_name: Optional[str] = None) -> List[Dict]:
        """
        List all files with optional filtering.
        
        Args:
            company_name: Filter by company name (optional)
        
        Returns:
            List of file metadata
        """
        files = list(self.metadata.values())
        
        if company_name:
            files = [f for f in files if f['company_name'].upper() == company_name.upper()]
        
        # Sort by creation date (newest first)
        files.sort(key=lambda x: x['created_at'], reverse=True)
        
        return files
    
    def delete_file(self, file_id: str) -> bool:
        """
        Delete file and its metadata.
        
        Args:
            file_id: File ID
        
        Returns:
            True if successful, False otherwise
        """
        if file_id in self.metadata:
            try:
                # Delete physical file
                stored_path = Path(self.metadata[file_id]['stored_path'])
                if stored_path.exists():
                    stored_path.unlink()
                
                # Remove metadata
                del self.metadata[file_id]
                self._save_metadata()
                
                _log.info(f"File deleted: {file_id}")
                return True
            except Exception as e:
                _log.error(f"Error deleting file: {e}")
                return False
        
        return False
    
    def cleanup_old_files(self, days: int = 30) -> int:
        """
        Clean up files older than specified days.
        
        Args:
            days: Delete files older than this many days
        
        Returns:
            Number of files deleted
        """
        from datetime import timedelta
        
        cutoff_date = datetime.now() - timedelta(days=days)
        deleted_count = 0
        
        files_to_delete = []
        for file_id, metadata in self.metadata.items():
            created_at = datetime.fromisoformat(metadata['created_at'])
            if created_at < cutoff_date:
                files_to_delete.append(file_id)
        
        for file_id in files_to_delete:
            if self.delete_file(file_id):
                deleted_count += 1
        
        _log.info(f"Cleaned up {deleted_count} old files")
        return deleted_count
