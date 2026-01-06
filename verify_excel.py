"""
Generate Excel and verify all fields are populated
"""
import json
from pathlib import Path
import openpyxl
from excel_generator import FinancialExcelGenerator

# Load JSON
json_path = Path("output/BRITANNIA_Britannia_Unaudited_Q2_June_2026/Britannia_Unaudited_Q2_June_2026-financial-data.json")
with open(json_path, 'r') as f:
    data = json.load(f)

print("="*80)
print("EXCEL GENERATION & VERIFICATION")
print("="*80)

# Generate Excel
output_path = Path("output/test_verification.xlsx")
generator = FinancialExcelGenerator()

print(f"\nGenerating Excel: {output_path}")
success = generator.generate_excel(data, output_path)

if success:
    print("✅ Excel generated successfully\n")
    
    # Load and verify
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # Check critical fields
    fields_to_check = [
        (8, 'Total Revenue', 'revenue_from_operations', ['B', 'D', 'G', 'C']),
        (9, 'Other income', 'other_income', ['B', 'D', 'G', 'C']),
        (10, 'Total Income', 'total_income', ['B', 'D', 'G', 'C']),
        (18, 'Purchases of stock-in-trade', 'purchases_stock_in_trade', ['B', 'D', 'G', 'C']),
        (22, 'Depreciation and amortisation', 'depreciation_amortisation_expense', ['B', 'D', 'G', 'C']),
        (30, 'PBT before exp items', 'profit_before_exceptional_and_tax', ['B', 'D', 'G', 'C']),
        (39, 'Deferred tax', 'deferred_tax', ['B', 'D', 'G', 'C']),
    ]
    
    print(f"{'Row':<5} | {'Field Name':<30} | {'Column':<8} | {'Value':<12} | {'Expected':<12} | Status")
    print("-"*100)
    
    all_match = True
    
    # Map columns to periods
    col_to_period = {
        'B': '30.06.2025',
        'C': '31.03.2025_Y',
        'D': '31.03.2025',
        'E': '31.12.2024',
        'F': '30.09.2024',
        'G': '30.06.2024',
    }
    
    # Get JSON values for comparison
    json_map = {item['key']: item['values'] for item in data['financial_data']}
    
    for row_num, field_name, json_key, cols in fields_to_check:
        label = ws[f'A{row_num}'].value
        for col in cols:
            if col in col_to_period:
                period = col_to_period[col]
                excel_value = ws[f'{col}{row_num}'].value
                expected_value = json_map.get(json_key, {}).get(period, '')
                
                # Parse values for comparison
                if excel_value == '-':
                    excel_value_str = '(empty)'
                elif excel_value is None:
                    excel_value_str = '(none)'
                else:
                    excel_value_str = str(excel_value)
                
                if expected_value == '':
                    expected_value_str = '(empty)'
                else:
                    expected_value_str = expected_value
                
                # Check match
                match = False
                if excel_value == '-' and expected_value == '':
                    match = True
                elif excel_value is not None and expected_value:
                    # Compare numeric values
                    try:
                        excel_numeric = float(str(excel_value).replace(',', '').replace('(', '-').replace(')', ''))
                        expected_numeric = float(expected_value.replace(',', '').replace('(', '-').replace(')', ''))
                        match = abs(excel_numeric - expected_numeric) < 0.01
                    except:
                        match = str(excel_value) == expected_value
                
                status = "✅" if match else "❌"
                if not match:
                    all_match = False
                
                print(f"{row_num:<5} | {field_name[:30]:<30} | {col}{row_num:<7} | {excel_value_str:<12} | {expected_value_str:<12} | {status}")
    
    print("\n" + "="*80)
    if all_match:
        print("✅ ALL FIELDS VERIFIED - Data correctly populated in Excel!")
    else:
        print("❌ SOME FIELDS MISSING - Data not correctly populated")
    
    # Check which periods have data in Excel columns
    print("\n" + "="*80)
    print("Period columns in Excel:")
    print("-"*80)
    for col in 'BCDEFGHIJKL':
        header = ws[f'{col}2'].value
        desc = ws[f'{col}3'].value
        # Count non-empty values in column
        non_empty = sum(1 for row in range(4, 50) if ws[f'{col}{row}'].value and ws[f'{col}{row}'].value != '-')
        status = "✓ HAS DATA" if non_empty > 0 else "○ EMPTY"
        print(f"  Column {col}: {header:20} | {desc:25} | {non_empty:2} values | {status}")
    
else:
    print("❌ Excel generation failed")

print("="*80)
